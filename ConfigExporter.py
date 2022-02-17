import argparse
import json
import os
import os.path
import re
import sys
import time
import zipfile
import zlib

import json_minify
import openpyxl.worksheet.worksheet

import CmdColorUtil
import common_util
from monkey_xls import *

"""
配置导出工具，支持导出类结构和外载json，其中类结构还支持自定义模板，通过自定义模板结构体，可以兼容多种语言
"""

# 模板配置文件 0template.json
template_config = None
cfg_vo_map = {}
file_count = 0
verbose = 0

# 生成配置结构体vo
OP_STRUCT = 0b1
# 生成json数据
OP_PACK = 0b10

app_dir = None


def error(content):
    CmdColorUtil.printRed('[ERROR] ' + content)


def warning(content):
    CmdColorUtil.printYellow('[WARNING] ' + content)


# 通过key获取模板配置数据
def get_cfg_by_key(p_key) -> TempCfgVo:
    global cfg_vo_map
    path0 = app_dir / 'template/0template.json'
    if p_key not in cfg_vo_map:
        global template_config

        if not template_config:
            # 加载模板配置
            with open(str(path0), 'r', encoding='utf-8') as f:
                # json_minify库支持json文件里面添加注释
                template_config = json.loads(json_minify.json_minify(f.read()))
                print('====加载模板文件配置成功\n{0}'.format(path0.resolve().absolute()))

        if p_key not in template_config:
            error('0template.json 中不存在 ' + p_key + ' 配置：')
            exit()

        if 'base' in template_config:  # 有基础配置先设置基础配置
            cfg_vo = TempCfgVo(template_config['base'])
            cfg_vo.set_data(template_config[p_key])
        else:
            cfg_vo = TempCfgVo(template_config[p_key])
        cfg_vo.app_dir = app_dir  # 设置工具的目录
        cfg_vo_map[p_key] = cfg_vo
        path_tmp: Path = path0.parent / cfg_vo.template
        if path_tmp.exists():
            print('====成功加载类结构模板\n{0}\n'.format(path_tmp.resolve().absolute()))
        else:
            error('...[warning]类结构模板不存在\n{0}\n'.format(path_tmp.resolve().absolute()))
    return cfg_vo_map[p_key]


def export_config_struct(excel_vo: ExcelVo, p_str_map):
    vo_list = excel_vo.key_vo_list
    # 跳过无key列表的数据列表
    if len(vo_list) == 0:
        return

    # 正则表达式官方文档参考
    # https://docs.python.org/zh-cn/3.7/library/re.html
    # re.M 让$生效
    # re.DOTALL 让.可以匹配换行符
    def rpl_loop(m):
        result = ''
        loop_str = str(m.group(1)).lstrip('\n')
        for v in vo_list:
            if not v.key_client:  # 跳过没有导出前端的字段
                continue

            def rpl_property(m1):
                key_str = m1.group(1)
                return replace_key(key_str, p_excel_vo=excel_vo, p_key_vo=v)

            result += re.sub('<#(.*?)#>', rpl_property, loop_str)
        # 这里需要把前后的换行干掉
        return result.rstrip('\n')

    output_str = re.sub('<<<<\\s*?$(.+?)>>>>\\s*?$', rpl_loop, excel_vo.cfg.str_tmp, flags=re.M | re.DOTALL)

    def rpl_export(m):
        key_str = m.group(1)
        return replace_key(key_str, p_excel_vo=excel_vo)

    output_str = re.sub('<#(.*?)#>', rpl_export, output_str)

    if not excel_vo.cfg.struct_in_one:
        path = os.path.join(excel_vo.cfg.output_path, excel_vo.export_filename)
        root = os.path.dirname(path)
        if not os.path.exists(root):
            os.makedirs(root)  # 递归创建文件夹
        with open(path, 'w', encoding='utf-8') as f:
            f.write(output_str)
            if verbose:
                print('...成功导出', excel_vo.export_filename)
    if p_str_map is not None:
        p_str_map.append(output_str)
    return output_str


# 替换关键字
def replace_key(p_key: str, p_excel_vo: ExcelVo = None, p_key_vo: KeyVo = None, p_export_name: str = None,
                p_enum_class_name: str = None):
    key_name = p_key
    if '|' in p_key:  # 这里关键词支持参数模式
        key_list = p_key.split('|')
        key_name = key_list[0]
        param = key_list[1]
        obj_par = parse_param(param)

    if key_name == 'source_filename':
        return p_excel_vo.source_filename

    elif key_name == 'sheet_name':
        return p_excel_vo.sheet.title

    elif key_name == 'export_name':
        if p_excel_vo:
            return p_excel_vo.export_name
        else:
            return p_export_name

    elif key_name == 'export_class_name':
        return p_excel_vo.export_class_name

    elif p_key_vo is not None:
        if key_name == 'property_name':
            return p_key_vo.key_client
        elif key_name == 'type':
            return transform_tye(p_key_vo.type, p_excel_vo.cfg.type_map)
        elif key_name == 'comment':
            sep = ''
            if obj_par:
                if 'separator' in obj_par:
                    # 批注分隔符
                    sep = obj_par['separator']
            comment = p_key_vo.comment
            if p_key_vo.pz:  # 批注功能
                if sep:
                    result = re.findall('^.*$', p_key_vo.pz, re.M)
                    for pzv in result:
                        comment += '\n' + sep + pzv
                else:
                    comment += '\n' + p_key_vo.pz
            return comment
        elif key_name == 'index':
            # 这里需要注意，python的数字类型不会自动转换为字符串，这里需要强转一下
            return str(p_key_vo.index)

    elif key_name == 'enum_class_name':
        return p_enum_class_name

    else:
        return 'undefined'


def parse_param(p_param: str):
    obj = {}
    params = p_param.split(',')
    for v in params:
        kvs = v.split(':')
        key = kvs[0]
        value = kvs[1]
        obj[key] = value
    return obj


# 根据配置转换类型
def transform_tye(p_type, p_map):
    if p_type in p_map:
        return p_map[p_type]
    else:
        if (p_type == '' or p_type is None) and p_map['default']:
            # 配置中typeMap字段有添default的则使用默认的类型
            return p_map['default']
        return None


def is_int(p_type, p_map):
    """
    判断是否整型
    :param p_type:
    :param p_map:
    :return:
    """
    ttype = transform_tye(p_type, p_map)
    if ttype == 'number':
        # 整型的类型可以在这里添加
        return True
    else:
        return False


# def parse_value(kv: KeyVo, cell: xlrd.sheet.Cell, p_map):
def parse_value(kv: KeyVo, cell: Cell, p_map):
    if is_int(kv.type, p_map):  # 整型
        if cell.value == '' or cell.value is None:  # 空值则强转为0
            value = 0
        else:
            value = int(cell.value)
    else:
        if cell.value is None:
            value = ''
        else:
            if cell.data_type == opencell.TYPE_NUMERIC:
                if cell.value % 1 == 0.0:
                    value = str(int(cell.value))
                else:
                    value = str(cell.value)
            else:
                value = str(cell.value)
    return value


# 导出配置数据文件
def export_json_data(excel_vo: ExcelVo, json_map):
    sheet = excel_vo.sheet
    key_vo_list = excel_vo.key_vo_list
    obj_list = {}
    for i in range(ExcelIndexEnum.data_start_r.value, sheet.max_row):
        # rows = sheet.row(i)
        rows = sheet[i + 1]  # openpyxl的row和col起始是1
        obj = {}
        ok_flag = True
        for v in key_vo_list:
            if not v.key_client:  # 跳过没有key名的
                continue
            cell: Cell = rows[v.index]
            # 跳过id为空的行
            if v.index == ExcelIndexEnum.data_start_c.value and cell.value is None:
                ok_flag = False
                break

            try:
                value = parse_value(v, cell, excel_vo.cfg.type_map)
            except BaseException as err:
                col_num = v.index + 1
                col_num_str = common_util.covert_10_to_26(col_num)
                export_name = excel_vo.export_name + '.json'
                error(
                    '{0} | sheet:{5} | {4} | 表格数值类型解析错误，请检查 {1}行 {2}({3})列'.format(excel_vo.source_filename,
                                                                                   i + 1,
                                                                                   col_num,
                                                                                   col_num_str,
                                                                                   export_name,
                                                                                   excel_vo.sheet.title))
                value = cell.value
            obj[v.key_client] = value
        if ok_flag:
            # obj_list.append(obj)
            if obj['id'] in obj_list:
                export_name = excel_vo.export_name + '.json'
                warning('{0} | sheet:{1} | {2} | 重复的id，原来的值会被覆盖，请检查 {3}行'.format(excel_vo.source_filename,
                                                                                 excel_vo.sheet.title,
                                                                                 export_name,
                                                                                 i + 1,
                                                                                 ))
            obj_list[obj['id']] = obj
        # print(obj)
    json_map[excel_vo.export_name] = obj_list

    # 散文件输出
    if not excel_vo.cfg.json_pack_in_one:
        # 未格式化的json
        json_obj_min = json.dumps(obj_list, ensure_ascii=False, separators=(',', ':'))
        path = os.path.join(excel_vo.cfg.json_path, excel_vo.export_name + '.json')
        root = os.path.dirname(path)
        if not os.path.exists(root):
            os.makedirs(root)  # 递归创建文件夹
        with open(path, 'w', encoding='utf-8') as f:
            f.write(json_obj_min)

        if excel_vo.cfg.json_compress == 'zlib' or excel_vo.cfg.json_compress == 'zip':
            zlib_path = os.path.join(excel_vo.cfg.json_path, excel_vo.export_name + '.' + excel_vo.cfg.compress_suffix)
            file_compress(path, zlib_path, delete_source=True, ptype=excel_vo.cfg.json_compress)

    if excel_vo.cfg.json_copy_path:
        # 格式化过的json（便于人员察看检查）
        json_obj_format = json.dumps(obj_list, indent=4, ensure_ascii=False)
        path = os.path.join(excel_vo.cfg.json_copy_path, excel_vo.export_name + '.json')
        root = os.path.dirname(path)
        if not os.path.exists(root):
            os.makedirs(root)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(json_obj_format)


# 导出vo文件
def main_run(p_key, op, p_verbose=0, p_source='', p_output='', p_json=''):
    cfg = get_cfg_by_key(p_key)
    global verbose
    verbose = p_verbose

    if p_source:
        cfg.source_path = p_source
    path_source = Path(cfg.source_path)
    path_source = path_source.resolve()
    if not path_source.exists():
        error('...[warning]路径不存在 {0}'.format(path_source))
        return

    if p_output:
        cfg.output_path = p_output
    path_output = Path(cfg.output_path)
    path_output = path_output.resolve()
    if not path_output.exists():
        error('...[warning]路径不存在 {0}'.format(path_output))
        return

    if p_json:
        cfg.json_path = p_json

    # 清除旧文件
    if cfg.clean:
        if ((op & OP_STRUCT) == OP_STRUCT) and cfg.output_path:
            for root, dirs, files in os.walk(cfg.output_path):
                for fname in files:
                    file_url = os.path.join(root, fname)
                    name, ext = os.path.splitext(file_url)
                    if ext == '.' + cfg.suffix:  # 删除指定格式的旧文件
                        os.remove(file_url)
        if (op & OP_PACK) == OP_PACK:
            if cfg.json_copy_path:
                for root, dirs, files in os.walk(cfg.json_copy_path):
                    for fname in files:
                        file_url = os.path.join(root, fname)
                        name, ext = os.path.splitext(file_url)
                        if ext == '.json':
                            os.remove(file_url)
            if cfg.json_path:
                for root, dirs, files in os.walk(cfg.json_path):
                    for fname in files:
                        file_url = os.path.join(root, fname)
                        name, ext = os.path.splitext(file_url)
                        if ext == '.json' or ext == cfg.compress_suffix:
                            os.remove(file_url)

    global file_count
    json_map = {}
    str_map = []
    export_name_map = {}
    # 遍历文件夹内所有的xlsx文件
    list_file = sorted(path_source.rglob('*.xlsx'))
    for v in list_file:
        file_url = v.absolute()
        if v.name.find('~$') > -1:  # 跳过临时打开xlsx文件
            continue
        # 使用data_only选项，我们从单元格而不是公式中获取值
        wb = openpyxl.load_workbook(filename=file_url, data_only=True)
        sheet_names = wb.sheetnames
        for i in sheet_names:
            sheet: openpyxl.worksheet.worksheet.Worksheet = wb[i]
            cell_first = sheet.cell(1, 1)  # 注意：openpyxl使用index从1开始，区别于xlrd
            if cell_first.value is None:  # 跳过第一行第一列没有数据的表
                break
            if p_verbose:
                print('{0} | sheet:{1}'.format(v.absolute(), sheet.title))
            excel_vo = ExcelVo(cfg=cfg, sheet=sheet, source_path=file_url, filename=v.name)
            if excel_vo.export_name in export_name_map:
                warning('...[warning]导出表名重复，跳过 {0} | sheet:{1}'.format(v, sheet.title))
                break
            if not excel_vo.has_id_in_client():  # 跳过没有id字段的
                warning('...[warning]缺少id字段，跳过 {0} | sheet:{1}'.format(v, sheet.title))
                break
            export_name_map[excel_vo.export_name] = excel_vo
            file_count += 1
            if (op & OP_STRUCT) == OP_STRUCT:  # 导出vo类
                export_config_struct(excel_vo, str_map)
            if (op & OP_PACK) == OP_PACK:  # 导出json数据
                export_json_data(excel_vo, json_map)

    # 导出枚举类
    if ((op & OP_STRUCT) == OP_STRUCT) and cfg.enum_tmp and cfg.enum_class_name:
        while True:
            enum_tmp = app_dir / 'template' / cfg.enum_tmp
            if not enum_tmp.exists():
                error('...[warning]不存在模板文件 {0}'.format(enum_tmp))
                break
            str_tmp = enum_tmp.read_text(encoding='utf-8')

            def rpl_loop(m):
                result = ''
                loop_str = str(m.group(1)).lstrip('\n')
                for k, v in export_name_map.items():
                    def rpl_property(m1):
                        key_str = m1.group(1)
                        return replace_key(key_str, p_excel_vo=v, p_export_name=k)

                    result += re.sub('<#(.*?)#>', rpl_property, loop_str)
                # 这里需要把前后的换行干掉
                return result.rstrip('\n')

            output_str = re.sub('<<<<\\s*?$(.+?)>>>>\\s*?$', rpl_loop, str_tmp, flags=re.M | re.DOTALL)

            def rpl_export(m):
                key_str = m.group(1)
                return replace_key(key_str, p_enum_class_name=cfg.enum_class_name)

            output_str = re.sub('<#(.*?)#>', rpl_export, output_str)
            path_enum = Path(cfg.output_path) / (cfg.enum_class_name + '.' + cfg.suffix)
            path_enum.parent.mkdir(parents=True, exist_ok=True)
            path_enum.write_text(output_str, encoding='utf-8')
            break

    # 所有配置打包到一个文件中
    if (op & OP_PACK) == OP_PACK and cfg.json_pack_in_one:
        json_pack = json.dumps(json_map, ensure_ascii=False, separators=(',', ':'))
        path = os.path.join(cfg.json_path, '0config' + '.json')
        root = os.path.dirname(path)
        if not os.path.exists(root):
            os.makedirs(root)  # 递归创建文件夹
        with open(path, 'w', encoding='utf-8') as f:
            f.write(json_pack)

        if cfg.json_compress == 'zlib' or cfg.json_compress == 'zip':  # zlib压缩 or zip打包
            zlib_path = os.path.join(cfg.json_path, '0config' + '.' + cfg.compress_suffix)
            file_compress(path, zlib_path, delete_source=True, ptype=cfg.json_compress)

        if cfg.json_copy_path:
            json_pack = json.dumps(json_map, ensure_ascii=False, indent=4)
            path = os.path.join(cfg.json_copy_path, '0config' + '.json')
            root = os.path.dirname(path)
            if not os.path.exists(root):
                os.makedirs(root)  # 递归创建文件夹
            with open(path, 'w', encoding='utf-8') as f:
                f.write(json_pack)
    if (op & OP_STRUCT) == OP_STRUCT and cfg.struct_in_one:
        if len(str_map) > 0:
            struct_str = ''
            for v in str_map:
                struct_str += v + '\n'
            path_struct: Path = Path(cfg.output_path) / ('ConfigStruct.' + cfg.suffix)
            path_struct.parent.mkdir(parents=True, exist_ok=True)
            path_struct.write_text(struct_str, encoding='utf-8')


def file_compress(spath, tpath, level=9, delete_source=False, ptype='zlib'):
    """
    zlib.compressobj 用来压缩数据流，用于文件传输
    :param spath:源文件
    :param tpath:目标文件
    :param level:压缩等级，越高压缩率越大，对应解压时间也变大
    :param delete_source:是否删除源文件，默认不删除
    :return:
    """
    if ptype == 'zlib':
        file_source = open(spath, 'rb')
        file_target = open(tpath, 'wb')
        compress_obj = zlib.compressobj(level, wbits=-15, method=zlib.DEFLATED)  # 压缩对象
        data = file_source.read(1024)  # 1024为读取的size参数
        while data:
            file_target.write(compress_obj.compress(data))  # 写入压缩数据
            data = file_source.read(1024)  # 继续读取文件中的下一个size的内容
        file_target.write(compress_obj.flush())  # compressobj.flush()包含剩余压缩输出的字节对象，将剩余的字节内容写入到目标文件中
        file_source.close()
        file_target.close()
    elif ptype == 'zip':
        try:
            with zipfile.ZipFile(tpath, mode="w", compression=zipfile.ZIP_DEFLATED) as f:
                path_source = Path(spath)
                f.write(spath, arcname=path_source.name)  # 写入压缩文件，会把压缩文件中的原有覆盖
        except Exception as e:
            print("异常对象的类型是:%s" % type(e))
            print("异常对象的内容是:%s" % e)
        finally:
            f.close()
        pass

    # 删除压缩源
    if delete_source:
        os.remove(spath)


def file_decompress(spath, tpath):
    """
    解压文件
    :param spath:源文件
    :param tpath:目标文件
    :return:
    """
    file_source = open(spath, 'rb')
    file_target = open(tpath, 'wb')
    decompress_obj = zlib.decompressobj(wbits=-15)
    data = file_source.read(1024)
    while data:
        file_target.write(decompress_obj.decompress(data))
        data = file_source.read(1024)
    file_target.write(decompress_obj.flush())
    file_source.close()
    file_target.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='帮助信息')
    parser.add_argument('--template', type=str, default='ts', help='0template.json中配置的模板键名')
    parser.add_argument('--exportJson', type=int, default=1, help='是否导出json，1输出，0不输出，默认为1')
    parser.add_argument('--exportStruct', type=int, default=1, help='是否导出语言结构体，1输出，0不输出，默认为1')
    parser.add_argument('--verbose', type=int, default=0, help='输出详细信息，0不输出，1输出，默认为0')
    parser.add_argument('--source', type=str, default='', help='配置路径')
    parser.add_argument('--output', type=str, default='', help='类结构体路径')
    parser.add_argument('--json', type=str, default='', help='配置导出json路径')

    args = parser.parse_args()

    print('运行参数：\n--加载模板：{0} \n--是否导出json：{1} \n--是否导出结构体：{2}'.format(
        args.template,
        args.exportJson,
        args.exportStruct,
        args.verbose))

    op_value = 0
    if args.exportJson:
        op_value |= OP_PACK
    if args.exportStruct:
        op_value |= OP_STRUCT

    if not op_value:
        print('无操作，退出')
        exit()

    source = args.source
    output = args.output
    json_data = args.json

    start = time.time()
    app_dir = Path(sys.argv[0]).parent
    # print(app_dir.resolve())
    main_run(args.template, op_value, p_verbose=args.verbose, p_source=source, p_output=output, p_json=json_data)
    # file_compress('./data/0config.json', './data/0config.zlib')
    # file_decompress('./data/0config.zlib', './data/fuck.json')
    end = time.time()
    print('输出 {0} 个文件'.format(file_count))
    print('总用时 {0} s'.format(round(end - start, 2)))
